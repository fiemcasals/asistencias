from django.http import HttpResponse, HttpResponseForbidden
from django.utils.timezone import localtime
from django.utils import timezone
from django.contrib.auth import get_user_model
from io import BytesIO
from openpyxl import Workbook

from ..models import (
    Diplomatura, Materia, Clase, Asistencia,
    ProfesorMateria, InscripcionDiplomatura, InscripcionMateria
)

def _dt(v):
    """Formatea datetimes/fechas a texto legible (local)."""
    if v is None:
        return ""
    if hasattr(v, "tzinfo"):
        return localtime(v).strftime("%Y-%m-%d %H:%M:%S")
    return v.strftime("%Y-%m-%d")

def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for r in rows:
        ws.append(r)
    _autosize(ws)

def exportar_xlsx(request):
    # Solo Coordinadores (3) o Administradores (5)
    if not request.user.is_authenticated or request.user.nivel not in (3, 5):
        return HttpResponseForbidden("No autorizado.")

    User = get_user_model()

    wb = Workbook()
    wb.remove(wb.active)

    # === Usuarios ===
    ws = wb.create_sheet("Usuarios")
    headers = [
        "id", "email", "first_name", "second_name", "last_name", "second_last_name",
        "dni", "nivel", "is_active", "date_joined", "last_login",
    ]
    rows = []
    for u in User.objects.all().order_by("last_name", "first_name"):
        rows.append([
            u.id, u.email, u.first_name, getattr(u, "second_name", ""),
            u.last_name, getattr(u, "second_last_name", ""),
            u.dni, u.nivel, u.is_active, _dt(u.date_joined), _dt(u.last_login)
        ])
    _write_sheet(ws, headers, rows)

    # === Diplomaturas ===
    ws = wb.create_sheet("Diplomaturas")
    headers = ["id", "nombre", "descripcion", "codigo", "creada_por_email", "coordinadores_emails"]
    rows = []
    for d in Diplomatura.objects.all().order_by("nombre"):
        coords = ", ".join(d.coordinadores.values_list("email", flat=True))
        rows.append([
            d.id, d.nombre, d.descripcion, d.codigo,
            d.creada_por.email if d.creada_por else "",
            coords
        ])
    _write_sheet(ws, headers, rows)

    # === Materias ===
    ws = wb.create_sheet("Materias")
    headers = ["id", "diplomatura_id", "diplomatura", "nombre", "descripcion", "codigo",
               "profesor_titular_id", "profesor_titular_email"]
    rows = []
    for m in Materia.objects.select_related("diplomatura", "profesor_titular").all().order_by("diplomatura__nombre", "nombre"):
        rows.append([
            m.id, m.diplomatura_id, m.diplomatura.nombre,
            m.nombre, m.descripcion, m.codigo,
            m.profesor_titular_id or "",
            m.profesor_titular.email if m.profesor_titular else ""
        ])
    _write_sheet(ws, headers, rows)

    # === Clases ===
    ws = wb.create_sheet("Clases")
    headers = ["id", "materia_id", "materia", "fecha", "hora_inicio", "hora_fin", "tema", "ventana_activa"]
    rows = []
    now = timezone.now()
    for c in Clase.objects.select_related("materia", "materia__diplomatura").all().order_by("-fecha"):
        ventana_activa = (c.hora_inicio <= now <= c.hora_fin)
        rows.append([
            c.id, c.materia_id, f"{c.materia.nombre} ({c.materia.diplomatura.nombre})",
            _dt(c.fecha), _dt(c.hora_inicio), _dt(c.hora_fin), c.tema, ventana_activa
        ])
    _write_sheet(ws, headers, rows)

    # === Asistencias ===
    ws = wb.create_sheet("Asistencias")
    headers = ["id", "clase_id", "materia", "fecha_clase",
               "user_id", "email_user", "dni_user", "presente", "timestamp"]
    rows = []
    asist_qs = Asistencia.objects.select_related(
        "clase", "clase__materia", "clase__materia__diplomatura", "user"
    ).all().order_by("-timestamp")
    for a in asist_qs:
        rows.append([
            a.id, a.clase_id,
            f"{a.clase.materia.nombre} ({a.clase.materia.diplomatura.nombre})",
            _dt(a.clase.fecha),
            a.user_id, a.user.email, a.user.dni,
            a.presente, _dt(a.timestamp)
        ])
    _write_sheet(ws, headers, rows)

    # === ProfesorMateria ===
    ws = wb.create_sheet("ProfesorMateria")
    headers = ["id", "user_id", "email", "materia_id", "materia", "diplomatura", "rol"]
    rows = []
    for pm in ProfesorMateria.objects.select_related("user", "materia", "materia__diplomatura").all():
        rows.append([
            pm.id, pm.user_id, pm.user.email,
            pm.materia_id, pm.materia.nombre, pm.materia.diplomatura.nombre, pm.rol
        ])
    _write_sheet(ws, headers, rows)

    # === InscripcionDiplomatura ===
    ws = wb.create_sheet("InscDiplomatura")
    headers = ["id", "user_id", "email", "dni", "diplomatura_id", "diplomatura", "fecha"]
    rows = []
    for ins in InscripcionDiplomatura.objects.select_related("user", "diplomatura").all():
        rows.append([
            ins.id, ins.user_id, ins.user.email, ins.user.dni,
            ins.diplomatura_id, ins.diplomatura.nombre, _dt(ins.fecha)
        ])
    _write_sheet(ws, headers, rows)

    # === InscripcionMateria ===
    ws = wb.create_sheet("InscMateria")
    headers = ["id", "user_id", "email", "dni", "materia_id", "materia", "diplomatura", "fecha"]
    rows = []
    for ins in InscripcionMateria.objects.select_related("user", "materia", "materia__diplomatura").all():
        rows.append([
            ins.id, ins.user_id, ins.user.email, ins.user.dni,
            ins.materia_id, ins.materia.nombre, ins.materia.diplomatura.nombre, _dt(ins.fecha)
        ])
    _write_sheet(ws, headers, rows)

    # ⚠️ No se exportan tokens para niveles < 5
    # (Si quisieras incluirlos solo para admin, podrías hacer un if request.user.nivel == 5:)

    # Respuesta HTTP
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"asistencias_export_{localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
